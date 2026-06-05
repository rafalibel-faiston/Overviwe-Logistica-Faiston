from fastapi import FastAPI, HTTPException, Response, Cookie, UploadFile, File, BackgroundTasks
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, HTMLResponse
from pydantic import BaseModel
from typing import Optional, List
import psycopg2
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import date, timedelta, datetime
from calendar import monthrange
import os, hashlib, secrets, csv, io
from dotenv import load_dotenv
from pathlib import Path

try:
    import openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

load_dotenv()

app = FastAPI(title="Faiston Ops - API", version="1.0")

Path("static/css").mkdir(parents=True, exist_ok=True)
Path("static/js").mkdir(parents=True, exist_ok=True)

sessions = {}  # token -> {id, nome, perfil, last_seen, page}

def _touch_session(token: str, page: str = ""):
    if token and token in sessions:
        sessions[token]["last_seen"] = datetime.utcnow()
        if page:
            sessions[token]["page"] = page

def get_db():
    try:
        conn = psycopg2.connect(os.environ.get("DATABASE_URL"))
        cur = conn.cursor()
        cur.execute("SET TIME ZONE 'America/Sao_Paulo'")
        cur.close()
        conn.commit()
        return conn
    except Exception as e:
        print(f"Erro BD: {e}")
        return None

def hash_senha(senha):
    return hashlib.sha256(senha.encode()).hexdigest()

def get_session(token: str, page: str = ""):
    sess = sessions.get(token)
    if sess:
        sess["last_seen"] = datetime.utcnow()
        if page: sess["page"] = page
    return sess

def setup_banco():
    conn = get_db()
    if not conn: return
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id SERIAL PRIMARY KEY,
                usuario VARCHAR(50) UNIQUE NOT NULL,
                senha_hash VARCHAR(64) NOT NULL,
                nome VARCHAR(100) NOT NULL,
                perfil VARCHAR(20) NOT NULL DEFAULT 'funcionario',
                ativo BOOLEAN DEFAULT TRUE,
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS primeiro_acesso BOOLEAN DEFAULT FALSE")
        cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS email VARCHAR(200) DEFAULT ''")
        cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS time VARCHAR(50) DEFAULT 'Projetos'")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS tarefas (
                id SERIAL PRIMARY KEY,
                usuario_id INTEGER REFERENCES usuarios(id),
                descricao TEXT NOT NULL,
                cliente VARCHAR(50),
                prioridade VARCHAR(20) DEFAULT 'Media',
                status VARCHAR(30) DEFAULT 'aberto',
                segundos INTEGER DEFAULT 0,
                criado_em TIMESTAMP DEFAULT NOW(),
                atualizado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS comentarios (
                id SERIAL PRIMARY KEY,
                tarefa_id INTEGER REFERENCES tarefas(id) ON DELETE CASCADE,
                usuario_id INTEGER REFERENCES usuarios(id),
                texto TEXT NOT NULL,
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS notificacoes (
                id SERIAL PRIMARY KEY,
                tipo VARCHAR(50) NOT NULL,
                mensagem TEXT NOT NULL,
                lida BOOLEAN DEFAULT FALSE,
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("SELECT id FROM usuarios WHERE usuario = 'admin'")
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO usuarios (usuario, senha_hash, nome, perfil) VALUES (%s, %s, %s, %s)",
                ('admin', hash_senha('admin123'), 'Administrador', 'admin')
            )
        cur.execute("""
            CREATE TABLE IF NOT EXISTS configuracoes (
                chave VARCHAR(100) PRIMARY KEY,
                valor TEXT NOT NULL,
                atualizado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit(); cur.close(); conn.close()
        print("✅ Banco configurado")
    except Exception as e:
        print(f"Erro setup: {e}")

setup_banco()

# --- MODELOS ---
class LoginRequest(BaseModel):
    usuario: str
    senha: str

TIMES_VALIDOS = ['Projetos', 'Logística', 'Rede Credenciada']

class NovoUsuario(BaseModel):
    usuario: str
    senha: str
    nome: str
    perfil: str
    email: str = ""
    time: str = "Projetos"

class AtualizarUsuario(BaseModel):
    nome: str
    perfil: str
    senha: str = ""
    ativo: bool = True
    email: str = ""
    time: str = "Projetos"

class TarefaModel(BaseModel):
    descricao: str
    cliente: str
    prioridade: str = "Media"
    status: str = "aberto"
    segundos: int = 0
    funcionario_id: Optional[int] = None
    projeto_id: Optional[int] = None
    data_prazo: Optional[str] = None
    data_agendamento: Optional[str] = None

class AtualizarSegundos(BaseModel):
    segundos: int

class TrocarSenhaModel(BaseModel):
    nova_senha: str

class AcaoBackoffice(BaseModel):
    comando: str
    cliente: str = "Geral"

# --- EMAIL ---
def enviar_email_acesso(destinatario: str, nome: str, usuario: str, senha) -> bool:
    system_url = os.environ.get("SYSTEM_URL", "https://dashboard-faiston-production.up.railway.app").rstrip("/")
    if not destinatario:
        return False
    try:
        perfil_map = {"admin": "Admin", "gestor": "Gestor", "funcionario": "Funcionário"}
        btn = f"<a href='{system_url}' style='display:block;background:linear-gradient(135deg,#5B2EE0,#B826C9);color:white;text-decoration:none;text-align:center;padding:14px;border-radius:10px;font-weight:700;font-size:14px;margin-bottom:12px'>Acessar o Sistema</a>"
        btn_ajuda = f"<a href='{system_url}/ajuda' style='display:block;background:white;color:#5B2EE0;text-decoration:none;text-align:center;padding:13px;border-radius:10px;font-weight:700;font-size:14px;margin-bottom:24px;border:2px solid #5B2EE0'>📖 Ver Guia de Uso</a>"
        ajuda = f"<a href='{system_url}/ajuda' style='color:#5B2EE0'>Guia de Uso</a>"
        if senha is None:
            bloco_senha = """<div style="background:#FFF8E6;border:1px solid #FFD166;border-radius:10px;padding:14px;margin-bottom:24px">
              <p style="color:#B8860B;font-size:13px;font-weight:700;margin:0 0 4px">🔑 Senha não alterada</p>
              <p style="color:#7A6020;font-size:13px;margin:0;line-height:1.5">Use a senha que você já cadastrou no sistema. Caso não lembre, entre em contato com o administrador para redefinir.</p>
            </div>"""
        else:
            bloco_senha = f"""<table width="100%" cellpadding="0" cellspacing="0" style="background:#F4F5FA;border-radius:12px;margin-bottom:24px">
              <tr><td style="padding:16px 20px 8px">
                <p style="color:#9097AC;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.8px;margin:0 0 4px">Usuário</p>
                <p style="color:#0B0D1F;font-size:16px;font-weight:700;font-family:monospace;margin:0">{usuario}</p>
              </td></tr>
              <tr><td style="padding:0 20px"><div style="height:1px;background:#E5E8F0"></div></td></tr>
              <tr><td style="padding:8px 20px 16px">
                <p style="color:#9097AC;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.8px;margin:0 0 4px">Senha temporária</p>
                <p style="color:#5B2EE0;font-size:20px;font-weight:900;font-family:monospace;letter-spacing:2px;margin:0">{senha}</p>
              </td></tr>
            </table>
            <div style="background:#FFF8E6;border:1px solid #FFD166;border-radius:10px;padding:14px;margin-bottom:24px">
              <p style="color:#B8860B;font-size:13px;font-weight:700;margin:0 0 4px">⚠️ Troca de senha obrigatória</p>
              <p style="color:#7A6020;font-size:13px;margin:0;line-height:1.5">No primeiro acesso, o sistema pedirá que você crie uma senha pessoal.</p>
            </div>"""
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;background:#F4F5FA;padding:32px 16px">
          <div style="background:linear-gradient(135deg,#5B2EE0,#B826C9,#EC4899);border-radius:16px;padding:32px;text-align:center;margin-bottom:24px">
            <h1 style="color:white;margin:0;font-size:26px;font-weight:900;letter-spacing:-0.5px">Faiston OPS</h1>
            <p style="color:rgba(255,255,255,0.8);margin:8px 0 0;font-size:14px">Torre de Controle</p>
          </div>
          <div style="background:white;border-radius:16px;border:1px solid #E5E8F0;padding:28px;margin-bottom:16px">
            <p style="color:#0B0D1F;font-size:17px;font-weight:700;margin:0 0 8px">Olá, {nome}!</p>
            <p style="color:#5E647A;font-size:14px;margin:0 0 24px;line-height:1.6">Seu acesso ao <strong>Faiston OPS</strong> foi criado. Abaixo estão suas credenciais de entrada:</p>
            {bloco_senha}
            {btn}
            {btn_ajuda}
            <div style="border-top:1px solid #E5E8F0;padding-top:20px">
              <p style="color:#9097AC;font-size:11px;margin:0 0 10px;font-weight:700;text-transform:uppercase;letter-spacing:0.5px">Primeiros passos</p>
              <ol style="color:#5E647A;font-size:13px;margin:0;padding-left:20px;line-height:2.2">
                <li>Acesse o sistema com o usuário e senha acima</li>
                <li>Crie sua senha pessoal (mín. 6 caracteres)</li>
                <li>Explore o <strong>Guia de Uso</strong> em {ajuda} para conhecer o sistema</li>
              </ol>
            </div>
          </div>
          <p style="color:#9097AC;font-size:11px;text-align:center;margin:0">Faiston OPS · Este é um email automático, não responda.</p>
        </div>
        """
        subject = "Seu acesso ao Faiston OPS"

        import urllib.request, json as _json

        # Brevo (API HTTP — funciona no Railway)
        brevo_key = os.environ.get("BREVO_API_KEY", "")
        email_user = os.environ.get("EMAIL_USER", "")
        if brevo_key and email_user:
            payload = _json.dumps({
                "sender": {"name": "Faiston OPS", "email": email_user},
                "to": [{"email": destinatario}],
                "subject": subject,
                "htmlContent": html,
            }).encode()
            req = urllib.request.Request(
                "https://api.brevo.com/v3/smtp/email",
                data=payload,
                headers={"api-key": brevo_key, "Content-Type": "application/json"},
                method="POST"
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                result = _json.loads(resp.read())
            print(f"[email-acesso] Brevo OK — id {result.get('messageId')} → {destinatario}")
            return True

        print("[email-acesso] Nenhuma configuração de email encontrada (BREVO_API_KEY + EMAIL_USER)")
        return False
    except Exception as e:
        print(f"[email-acesso] Falha ao enviar para {destinatario}: {e}")
        return False

# --- AUTH ---
@app.post("/api/login")
def login(req: LoginRequest, response: Response):
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, nome, perfil, COALESCE(primeiro_acesso, FALSE), COALESCE(time,'Projetos') FROM usuarios WHERE usuario=%s AND senha_hash=%s AND ativo=TRUE",
                    (req.usuario, hash_senha(req.senha)))
        row = cur.fetchone()
        cur.close(); conn.close()
        if not row: raise HTTPException(status_code=401, detail="Usuário ou senha inválidos")
        token = secrets.token_hex(32)
        sessions[token] = {"id": row[0], "nome": row[1], "perfil": row[2], "time": row[4], "last_seen": datetime.utcnow(), "page": "dashboard"}
        response.set_cookie("faiston_token", token, httponly=True, samesite="lax", max_age=86400)
        return {"sucesso": True, "perfil": row[2], "nome": row[1], "primeiro_acesso": bool(row[3])}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/logout")
def logout(response: Response, faiston_token: str = Cookie(None)):
    if faiston_token and faiston_token in sessions: del sessions[faiston_token]
    response.delete_cookie("faiston_token")
    return {"sucesso": True}

@app.post("/api/trocar-senha")
def trocar_senha(body: TrocarSenhaModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    if not body.nova_senha or len(body.nova_senha) < 6:
        raise HTTPException(status_code=400, detail="A nova senha deve ter pelo menos 6 caracteres.")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("UPDATE usuarios SET senha_hash=%s, primeiro_acesso=FALSE WHERE id=%s",
                    (hash_senha(body.nova_senha), sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/me")
def me(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    return sess

# --- USUÁRIOS ---
@app.get("/api/usuarios")
def listar_usuarios(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403, detail="Acesso negado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        if sess["perfil"] == "admin":
            cur.execute("SELECT id, usuario, nome, perfil, ativo, criado_em, COALESCE(email,''), COALESCE(time,'Projetos') FROM usuarios WHERE ativo=TRUE ORDER BY criado_em DESC")
        else:
            cur.execute("SELECT id, usuario, nome, perfil, ativo, criado_em, COALESCE(email,''), COALESCE(time,'Projetos') FROM usuarios WHERE ativo=TRUE AND COALESCE(time,'Projetos')=%s ORDER BY criado_em DESC", (sess.get("time","Projetos"),))
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"id": r[0], "usuario": r[1], "nome": r[2], "perfil": r[3], "ativo": r[4], "criado_em": str(r[5]), "email": r[6], "time": r[7]} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/funcionarios")
def listar_funcionarios(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        if sess["perfil"] == "admin":
            cur.execute("SELECT id, nome FROM usuarios WHERE perfil='funcionario' AND ativo=TRUE ORDER BY nome")
        else:
            cur.execute("SELECT id, nome FROM usuarios WHERE perfil='funcionario' AND ativo=TRUE AND COALESCE(time,'Projetos')=%s ORDER BY nome", (sess.get("time","Projetos"),))
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"id": r[0], "nome": r[1]} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/usuarios")
def criar_usuario(u: NovoUsuario, bg: BackgroundTasks, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403, detail="Acesso negado")
    if u.perfil not in ("admin", "gestor", "funcionario", "demo"): raise HTTPException(status_code=400, detail="Perfil inválido")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        time_val = u.time if u.time in TIMES_VALIDOS else "Projetos"
        cur.execute("INSERT INTO usuarios (usuario, senha_hash, nome, perfil, email, primeiro_acesso, time) VALUES (%s, %s, %s, %s, %s, TRUE, %s) RETURNING id",
                    (u.usuario, hash_senha(u.senha), u.nome, u.perfil, u.email, time_val))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        tem_email = bool(u.email)
        if tem_email:
            bg.add_task(enviar_email_acesso, u.email, u.nome, u.usuario, u.senha)
        return {"sucesso": True, "id": new_id, "email_enviado": tem_email}
    except psycopg2.errors.UniqueViolation: raise HTTPException(status_code=400, detail="Usuário já existe")
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/usuarios/{uid}")
def atualizar_usuario(uid: int, u: AtualizarUsuario, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403, detail="Acesso negado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        time_val = u.time if u.time in TIMES_VALIDOS else "Projetos"
        if u.senha:
            cur.execute("UPDATE usuarios SET nome=%s, perfil=%s, ativo=%s, email=%s, senha_hash=%s, primeiro_acesso=TRUE, time=%s WHERE id=%s",
                        (u.nome, u.perfil, u.ativo, u.email, hash_senha(u.senha), time_val, uid))
        else:
            cur.execute("UPDATE usuarios SET nome=%s, perfil=%s, ativo=%s, email=%s, time=%s WHERE id=%s",
                        (u.nome, u.perfil, u.ativo, u.email, time_val, uid))
        conn.commit(); cur.close(); conn.close()
        # Atualiza sessões ativas deste usuário com o novo time/perfil sem precisar fazer logout
        for s in sessions.values():
            if s.get("id") == uid:
                s["time"] = time_val
                s["perfil"] = u.perfil
                s["nome"] = u.nome
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/usuarios/{uid}/reenviar-email")
def reenviar_email_acesso(uid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403, detail="Acesso negado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("SELECT nome, usuario, email FROM usuarios WHERE id=%s AND ativo=TRUE", (uid,))
        row = cur.fetchone(); cur.close(); conn.close()
        if not row: raise HTTPException(status_code=404, detail="Usuário não encontrado")
        nome, usuario, email = row
        if not email: raise HTTPException(status_code=400, detail="Este usuário não tem email cadastrado")
        enviado = enviar_email_acesso(email, nome, usuario, None)
        if not enviado: raise HTTPException(status_code=500, detail="Falha ao enviar email — verifique as variáveis EMAIL_USER e EMAIL_APP_PASSWORD no servidor")
        return {"sucesso": True}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/usuarios/{uid}")
def deletar_usuario(uid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403, detail="Acesso negado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM usuarios WHERE id=%s AND usuario != 'admin'", (uid,))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

# --- TAREFAS ---
@app.get("/api/tarefas")
def listar_tarefas(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        # Migração silenciosa: adiciona projeto_id se ainda não existe
        cur.execute("ALTER TABLE tarefas ADD COLUMN IF NOT EXISTS projeto_id INTEGER REFERENCES projetos(id)")
        cur.execute("ALTER TABLE tarefas ADD COLUMN IF NOT EXISTS data_prazo DATE")
        cur.execute("ALTER TABLE tarefas ADD COLUMN IF NOT EXISTS data_agendamento DATE")
        conn.commit()
        base_sel = """SELECT t.id, t.descricao, t.cliente, t.prioridade, t.status, t.segundos,
                             t.criado_em, u.nome, t.projeto_id, COALESCE(p.nome,'') AS projeto_nome,
                             t.data_prazo, t.data_agendamento
                      FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id
                      LEFT JOIN projetos p ON p.id = t.projeto_id"""
        if sess["perfil"] == "admin":
            cur.execute(base_sel + " ORDER BY t.criado_em DESC")
        elif sess["perfil"] in ("gestor", "demo"):
            cur.execute(base_sel + " WHERE COALESCE(u.time,'Projetos')=%s ORDER BY t.criado_em DESC", (sess.get("time","Projetos"),))
        else:
            cur.execute(base_sel + " WHERE t.usuario_id = %s ORDER BY t.criado_em DESC", (sess["id"],))
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"id": r[0], "descricao": r[1], "cliente": r[2], "prioridade": r[3],
                 "status": r[4], "segundos": r[5], "criado_em": str(r[6]), "funcionario": r[7],
                 "projeto_id": r[8], "projeto_nome": r[9],
                 "data_prazo": str(r[10]) if r[10] else None,
                 "data_agendamento": str(r[11]) if r[11] else None} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/tarefas")
def criar_tarefa(t: TarefaModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        # Gestor/admin pode atribuir a outro funcionário via funcionario_id
        uid = sess["id"]
        if t.funcionario_id and sess["perfil"] in ("admin", "gestor", "demo"):
            cur.execute("SELECT id FROM usuarios WHERE id=%s AND ativo=TRUE", (t.funcionario_id,))
            if cur.fetchone():
                uid = t.funcionario_id
        cur.execute(
            "INSERT INTO tarefas (usuario_id, descricao, cliente, prioridade, status, segundos, projeto_id, data_prazo, data_agendamento) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (uid, t.descricao, t.cliente, t.prioridade, t.status, t.segundos, t.projeto_id or None,
             t.data_prazo or None, t.data_agendamento or None)
        )
        new_id = cur.fetchone()[0]
        criar_notificacao(conn, "nova_tarefa", f"🆕 {sess['nome']} criou uma tarefa: {t.descricao[:50]} [{t.cliente}]")
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/tarefas/{tid}")
def atualizar_tarefa(tid: int, t: TarefaModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute(
            "UPDATE tarefas SET descricao=%s, cliente=%s, prioridade=%s, status=%s, segundos=%s, projeto_id=%s, data_prazo=%s, data_agendamento=%s, atualizado_em=NOW() WHERE id=%s AND usuario_id=%s",
            (t.descricao, t.cliente, t.prioridade, t.status, t.segundos, t.projeto_id or None,
             t.data_prazo or None, t.data_agendamento or None, tid, sess["id"])
        )
        if t.status == "concluido":
            criar_notificacao(conn, "tarefa_concluida", f"✅ {sess['nome']} concluiu: {t.descricao[:50]} [{t.cliente}]")
        elif t.status == "em_andamento":
            criar_notificacao(conn, "tarefa_iniciada", f"▶️ {sess['nome']} iniciou: {t.descricao[:50]} [{t.cliente}]")
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.patch("/api/tarefas/{tid}/segundos")
def atualizar_segundos(tid: int, body: AtualizarSegundos, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("UPDATE tarefas SET segundos=%s, atualizado_em=NOW() WHERE id=%s AND usuario_id=%s",
                    (body.segundos, tid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/tarefas/{tid}")
def deletar_tarefa(tid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        if sess["perfil"] in ("admin", "gestor", "demo"):
            cur.execute("DELETE FROM tarefas WHERE id=%s", (tid,))
        else:
            cur.execute("DELETE FROM tarefas WHERE id=%s AND usuario_id=%s", (tid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/admin/online")
def usuarios_online(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403)
    agora = datetime.utcnow()
    online = []
    for token, s in list(sessions.items()):
        last = s.get("last_seen")
        if last and (agora - last).total_seconds() < 300:  # ativo nos últimos 5 min
            minutos = int((agora - last).total_seconds() / 60)
            online.append({
                "nome": s["nome"],
                "perfil": s["perfil"],
                "page": s.get("page", "—"),
                "ultimo_acesso": f"há {minutos} min" if minutos > 0 else "agora"
            })
    return sorted(online, key=lambda x: x["ultimo_acesso"])

@app.get("/api/admin/atividades")
def atividades_recentes(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT u.nome, u.perfil, n.tipo, n.mensagem, n.criado_em
            FROM notificacoes n
            JOIN usuarios u ON n.usuario_id = u.id
            WHERE n.criado_em >= NOW() - INTERVAL '24 hours'
            ORDER BY n.criado_em DESC
            LIMIT 50
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"nome": r[0], "perfil": r[1], "tipo": r[2], "mensagem": r[3],
                 "quando": r[4].strftime("%H:%M")} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/admin/ping")
def ping_session(page: str = "", faiston_token: str = Cookie(None)):
    get_session(faiston_token, page)
    return {"ok": True}

@app.delete("/api/admin/limpar-tarefas")
def limpar_todas_tarefas(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403, detail="Apenas admin pode limpar tarefas")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM tarefas")
        deleted = cur.rowcount
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "removidas": deleted}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

# --- MÉTRICAS DASHBOARD ---
@app.get("/api/metricas")
def get_metricas(cliente: str = "", data_inicio: str = "", data_fim: str = "", funcionario: str = "", projeto: str = "", faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        conditions = []
        params = []
        # Filtro de time: gestor/demo vê só seu time; admin vê tudo
        is_admin = sess["perfil"] == "admin"
        if not is_admin:
            conditions.append("COALESCE(u.time,'Projetos') = %s")
            params.append(sess.get("time", "Projetos"))
        if cliente:
            conditions.append("t.cliente = %s")
            params.append(cliente)
        if funcionario:
            conditions.append("u.nome ILIKE %s")
            params.append(f"%{funcionario}%")
        if data_inicio:
            conditions.append("t.criado_em >= %s")
            params.append(data_inicio + " 00:00:00")
        if data_fim:
            conditions.append("t.criado_em <= %s")
            params.append(data_fim + " 23:59:59")
        if projeto:
            conditions.append("p.nome = %s")
            params.append(projeto)
        filtro = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        params = tuple(params)

        # JOINs necessários — time filter sempre precisa do JOIN com usuarios
        join_u = "JOIN usuarios u ON t.usuario_id = u.id" if (funcionario or not is_admin) else ""
        join_p = "LEFT JOIN projetos p ON t.projeto_id = p.id" if projeto else ""
        # Helpers para adicionar condição de status sem quebrar o filtro existente
        def fwhere(extra): return f"{filtro} AND {extra}" if filtro else f"WHERE {extra}"

        # KPIs gerais
        joins = f"{join_u} {join_p}"
        cur.execute(f"SELECT COUNT(*) FROM tarefas t {joins} {filtro}", params)
        total = cur.fetchone()[0]

        w_aberto = fwhere("t.status = 'aberto'")
        cur.execute(f"SELECT COUNT(*) FROM tarefas t {joins} {w_aberto}", params)
        abertos = cur.fetchone()[0]

        w_concluido = fwhere("t.status = 'concluido'")
        cur.execute(f"SELECT COUNT(*) FROM tarefas t {joins} {w_concluido}", params)
        concluidos = cur.fetchone()[0]

        w_andamento = fwhere("t.status = 'em_andamento'")
        cur.execute(f"SELECT COUNT(*) FROM tarefas t {joins} {w_andamento}", params)
        em_andamento = cur.fetchone()[0]

        cur.execute(f"SELECT COALESCE(SUM(t.segundos),0) FROM tarefas t {joins} {filtro}", params)
        total_segundos = cur.fetchone()[0]

        # Clientes ativos (distintos) — respeita filtro de time
        cur.execute(f"SELECT COUNT(DISTINCT t.cliente) FROM tarefas t {join_u} {join_p} {filtro}", params)
        clientes_ativos = cur.fetchone()[0]

        # Funcionários com tarefas — respeita filtro de time
        cur.execute(f"SELECT COUNT(DISTINCT t.usuario_id) FROM tarefas t {join_u} {join_p} {filtro}", params)
        funcionarios_ativos = cur.fetchone()[0]

        # SLA: % de tarefas concluídas sobre o total
        sla = round((concluidos / total * 100)) if total > 0 else 0

        # Média de horas por funcionário — respeita todos os filtros
        cur.execute(
            f"SELECT COUNT(DISTINCT t.usuario_id), COALESCE(SUM(t.segundos),0) FROM tarefas t {joins} {filtro}",
            params
        )
        row_media = cur.fetchone()
        n_funcs = max(row_media[0], 1)
        media_horas_func = round(row_media[1] / 3600 / n_funcs, 1)

        # Horas por cliente — respeita todos os filtros
        cur.execute(
            f"SELECT t.cliente, COALESCE(SUM(t.segundos),0) as total_seg "
            f"FROM tarefas t {joins} {filtro} GROUP BY t.cliente ORDER BY total_seg DESC",
            params
        )
        horas_por_cliente = [{"cliente": r[0], "horas": round(r[1]/3600, 1)} for r in cur.fetchall()]

        # Status da fila (para donut)
        cur.execute(f"SELECT t.status, COUNT(*) FROM tarefas t {joins} {filtro} GROUP BY t.status", params)
        status_fila = {r[0]: r[1] for r in cur.fetchall()}

        # Volume por dia da semana (últimos 7 dias — para área)
        cur.execute("""
            SELECT TO_CHAR(criado_em, 'Dy') as dia, COUNT(*) as total
            FROM tarefas
            WHERE criado_em >= NOW() - INTERVAL '7 days'
            GROUP BY TO_CHAR(criado_em, 'Dy'), DATE_TRUNC('day', criado_em)
            ORDER BY DATE_TRUNC('day', criado_em)
        """)
        volume_semana = [{"dia": r[0], "total": r[1]} for r in cur.fetchall()]

        # Funil de atendimento
        funil = {
            "Abertura": total,
            "Triagem": total - max(0, total - abertos),
            "Acionamento": em_andamento + concluidos,
            "Acompanhamento": em_andamento + concluidos,
            "Fechamento": concluidos
        }

        # Tarefas recentes
        q = f"""SELECT t.id, t.descricao, t.cliente, t.prioridade, t.status, t.segundos, t.criado_em, u.nome
            FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id {join_p}
            {filtro} ORDER BY t.criado_em DESC LIMIT 10"""
        cur.execute(q, params)
        recentes = [{"id": r[0], "descricao": r[1], "cliente": r[2], "prioridade": r[3],
                     "status": r[4], "segundos": r[5], "criado_em": str(r[6]), "funcionario": r[7]}
                    for r in cur.fetchall()]

        # Horas por funcionário — respeita todos os filtros incluindo time
        func_conds = list(conditions) + ["u.perfil = 'funcionario'"]
        if not is_admin and not any("u.time" in c for c in func_conds):
            func_conds.append("COALESCE(u.time,'Projetos') = %s")
            func_params = params + (sess.get("time", "Projetos"),)
        else:
            func_params = params
        func_filtro = "WHERE " + " AND ".join(func_conds)
        cur.execute(
            f"SELECT u.nome, COALESCE(SUM(t.segundos),0), COUNT(t.id) as total_tarefas "
            f"FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id {join_p} "
            f"{func_filtro} GROUP BY u.nome ORDER BY SUM(t.segundos) DESC",
            func_params
        )
        horas_por_func = [{"nome": r[0], "horas": round(r[1]/3600, 1), "tarefas": r[2]} for r in cur.fetchall()]

        # Taxa de conclusão por cliente — respeita todos os filtros
        cur.execute(
            f"SELECT t.cliente, COUNT(*) as total, "
            f"SUM(CASE WHEN t.status='concluido' THEN 1 ELSE 0 END) as concluidas "
            f"FROM tarefas t {joins} {filtro} GROUP BY t.cliente ORDER BY total DESC",
            params)
        taxa_rows = cur.fetchall()
        taxa_conclusao = [{"cliente": r[0], "total": r[1], "concluidas": r[2],
            "taxa": round(r[2]/r[1]*100) if r[1] > 0 else 0} for r in taxa_rows]

        cur.close(); conn.close()
        return {
            "kpis": {
                "total_tarefas": total,
                "tickets_abertos": abertos,
                "em_andamento": em_andamento,
                "concluidos": concluidos,
                "clientes_ativos": clientes_ativos,
                "funcionarios_ativos": funcionarios_ativos,
                "total_horas": round(total_segundos / 3600, 1),
                "sla": sla,
                "media_horas_func": media_horas_func
            },
            "horas_por_cliente": horas_por_cliente,
            "status_fila": status_fila,
            "volume_semana": volume_semana,
            "funil": funil,
            "recentes": recentes,
            "horas_por_func": horas_por_func,
            "taxa_conclusao": taxa_conclusao
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/registrar-acao")
def registrar_acao(acao: AcaoBackoffice, faiston_token: str = Cookie(None)):
    return {"sucesso": True, "mensagem": "Ação registrada"}

@app.get("/api/health")
def health(): return {"status": "ok"}

@app.get("/api/exportar")
def exportar_excel(cliente: str = "", data_inicio: str = "", data_fim: str = "", faiston_token: str = Cookie(None)):
    from fastapi.responses import StreamingResponse
    import io
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    if sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403, detail="Acesso negado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        conditions = []
        params = []
        if cliente:
            conditions.append("t.cliente = %s")
            params.append(cliente)
        if data_inicio:
            conditions.append("t.criado_em >= %s")
            params.append(data_inicio + " 00:00:00")
        if data_fim:
            conditions.append("t.criado_em <= %s")
            params.append(data_fim + " 23:59:59")
        filtro = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        cur.execute(f"""SELECT u.nome, t.descricao, t.cliente, t.prioridade, t.status,
            t.segundos, t.criado_em, t.atualizado_em
            FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id
            {filtro} ORDER BY t.criado_em DESC""", tuple(params))
        rows = cur.fetchall()
        cur.close(); conn.close()

        # Gerar XLSX com openpyxl
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tarefas"

        # Estilos
        header_fill = PatternFill("solid", fgColor="4A00E0")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alt_fill = PatternFill("solid", fgColor="F8FAFC")
        border = Border(bottom=Side(style='thin', color='E2E8F0'))
        center = Alignment(horizontal='center', vertical='center')

        # Cabeçalho
        headers = ["Funcionário", "Tarefa", "Cliente", "Prioridade", "Status", "Horas", "Minutos", "Total (h)", "Criado em", "Atualizado em"]
        col_widths = [25, 40, 20, 12, 15, 8, 8, 10, 18, 18]
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 30

        # Dados
        status_map = {"concluido": "Concluído", "em_andamento": "Em Andamento", "aberto": "Aberto"}
        prio_colors = {"Alta": "FFE4E6", "Media": "FEF3C7", "Baixa": "D1FAE5"}
        status_colors = {"concluido": "D1FAE5", "em_andamento": "CFFAFE", "aberto": "F1F5F9"}

        for i, r in enumerate(rows, 2):
            h = r[5] // 3600
            m = (r[5] % 3600) // 60
            total_h = round(r[5] / 3600, 2)
            status_label = status_map.get(r[4], r[4])
            row_data = [r[0], r[1], r[2], r[3], status_label, h, m, total_h,
                str(r[6])[:16] if r[6] else "", str(r[7])[:16] if r[7] else ""]
            fill = PatternFill("solid", fgColor="FFFFFF") if i % 2 == 0 else alt_fill
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                # Cor por prioridade e status
                if col == 4 and r[3] in prio_colors:
                    cell.fill = PatternFill("solid", fgColor=prio_colors[r[3]])
                elif col == 5 and r[4] in status_colors:
                    cell.fill = PatternFill("solid", fgColor=status_colors[r[4]])
                else:
                    cell.fill = fill
            ws.row_dimensions[i].height = 22

        # Totais
        total_row = len(rows) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=sum(r[5]//3600 for r in rows)).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=round(sum(r[5] for r in rows)/3600, 2)).font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"faiston_tarefas{'_'+cliente if cliente else ''}{'_'+data_inicio if data_inicio else ''}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/limpar-seed")
def limpar_seed(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403, detail="Apenas admin")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM tarefas WHERE usuario_id IN (SELECT id FROM usuarios WHERE usuario IN ('mariana','joao','carlos','fernanda','thiago'))")
        tarefas = cur.rowcount
        cur.execute("DELETE FROM tarefas WHERE descricao LIKE '%[TESTE]%'")
        tarefas += cur.rowcount
        cur.execute("DELETE FROM usuarios WHERE usuario IN ('mariana','joao','carlos','fernanda','thiago')")
        usuarios = cur.rowcount
        conn.commit()
        return {"ok": True, "tarefas_removidas": tarefas, "usuarios_removidos": usuarios}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/seed-dados")
def seed_dados():
    import random, hashlib
    from datetime import datetime, timedelta
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM tarefas WHERE descricao LIKE '%[TESTE]%'")
        cur.execute("DELETE FROM usuarios WHERE usuario IN ('mariana','joao','carlos','fernanda','thiago')")

        funcionarios = [
            ("mariana","faiston123","Mariana Silva","funcionario"),
            ("joao","faiston123","João Henrique","funcionario"),
            ("carlos","faiston123","Carlos Eduardo","funcionario"),
            ("fernanda","faiston123","Fernanda Lima","funcionario"),
            ("thiago","faiston123","Thiago Rocha","funcionario"),
        ]
        ids = {}
        for usuario, senha, nome, perfil in funcionarios:
            cur.execute("""INSERT INTO usuarios (usuario, senha_hash, nome, perfil, primeiro_acesso)
                VALUES (%s,%s,%s,%s,FALSE) ON CONFLICT (usuario) DO UPDATE SET nome=%s RETURNING id""",
                (usuario, hashlib.sha256(senha.encode()).hexdigest(), nome, perfil, nome))
            ids[nome] = cur.fetchone()[0]

        clientes = ["NTT","Arcos Dourados","Zamp","Telcoweb","VIVO VITA"]
        prioridades_peso = ["Critica","Alta","Alta","Media","Media","Media","Baixa"]
        descricoes = [
            "[TESTE] Abertura de chamado no NOC",
            "[TESTE] Acompanhamento de incidente crítico",
            "[TESTE] Configuração de switch core",
            "[TESTE] Monitoramento de links MPLS",
            "[TESTE] Troca de equipamento defeituoso",
            "[TESTE] Atualização de firmware",
            "[TESTE] Relatório de disponibilidade mensal",
            "[TESTE] Escalada para fornecedor",
            "[TESTE] Revisão de topologia de rede",
            "[TESTE] Acionamento de parceiro técnico",
            "[TESTE] Documentação de circuito",
            "[TESTE] Teste de failover",
            "[TESTE] Análise de log de erros",
            "[TESTE] Validação de SLA",
            "[TESTE] Suporte remoto ao cliente",
            "[TESTE] Instalação de CPE",
            "[TESTE] Diagnóstico de latência",
            "[TESTE] Follow-up de chamado crítico",
        ]
        now = datetime.now()
        total = 0

        # Mariana — sobrecarregada, muitos abertos antigos (IA deve alertar)
        for i in range(12):
            dias = random.randint(8, 20)
            cur.execute("INSERT INTO tarefas (usuario_id,descricao,cliente,prioridade,status,segundos,criado_em,atualizado_em) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                (ids["Mariana Silva"], random.choice(descricoes), random.choice(clientes),
                 random.choice(["Alta","Critica"]), "aberto", 0,
                 now - timedelta(days=dias), now - timedelta(days=dias)))
            total += 1

        # João — lento, tickets em andamento há muito tempo
        for i in range(8):
            dias = random.randint(5, 15)
            status = random.choice(["em_andamento","em_andamento","aberto"])
            cur.execute("INSERT INTO tarefas (usuario_id,descricao,cliente,prioridade,status,segundos,criado_em,atualizado_em) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                (ids["João Henrique"], random.choice(descricoes), random.choice(clientes),
                 random.choice(prioridades_peso), status, random.randint(600, 3600),
                 now - timedelta(days=dias), now - timedelta(days=dias)))
            total += 1

        # Carlos — equilibrado, maioria concluído
        for i in range(10):
            status = random.choice(["concluido","concluido","concluido","em_andamento","aberto"])
            segundos = random.randint(1800, 7200) if status == "concluido" else random.randint(600, 3600)
            cur.execute("INSERT INTO tarefas (usuario_id,descricao,cliente,prioridade,status,segundos,criado_em,atualizado_em) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                (ids["Carlos Eduardo"], random.choice(descricoes), random.choice(clientes),
                 random.choice(prioridades_peso), status, segundos,
                 now - timedelta(days=random.randint(1,7)), now))
            total += 1

        # Fernanda — poucos tickets, bem resolvidos
        for i in range(5):
            cur.execute("INSERT INTO tarefas (usuario_id,descricao,cliente,prioridade,status,segundos,criado_em,atualizado_em) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                (ids["Fernanda Lima"], random.choice(descricoes), random.choice(clientes),
                 "Media", random.choice(["concluido","concluido","aberto"]), random.randint(1800, 5400),
                 now - timedelta(days=random.randint(1,5)), now))
            total += 1

        # Thiago — vários críticos abertos
        for i in range(7):
            cur.execute("INSERT INTO tarefas (usuario_id,descricao,cliente,prioridade,status,segundos,criado_em,atualizado_em) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                (ids["Thiago Rocha"], random.choice(descricoes), random.choice(clientes),
                 random.choice(["Critica","Alta"]), random.choice(["aberto","aberto","em_andamento"]), 0,
                 now - timedelta(days=random.randint(2,10)), now))
            total += 1

        conn.commit(); cur.close(); conn.close()
        return {
            "sucesso": True,
            "tarefas_criadas": total,
            "usuarios": {u[0]: "senha: faiston123" for u in funcionarios}
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- PÁGINAS ---
@app.get("/")
def root(): return FileResponse("static/login.html")

@app.get("/dashboard")
def dashboard(): return FileResponse("static/index.html")

@app.get("/funcionario")
def funcionario(): return FileResponse("static/funcionario.html")

@app.get("/admin")
def admin_page(): return FileResponse("static/admin.html")

@app.get("/overview")
def overview_input_page(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: return FileResponse("static/login.html")
    return FileResponse("static/overview-input.html")

app.mount("/css", StaticFiles(directory="static/css"), name="css")
app.mount("/js", StaticFiles(directory="static/js"), name="js")

# ─── OVERVIEW SEMANAL ────────────────────────────────────────────────────────

AREAS_VALIDAS      = ['Logistica', 'Facilities', 'Seguros']
CATEGORIAS_VALIDAS = ['atividade','todo','criticidade','escalation','direcao','ideia','melhoria','chamado','proximo_passo']

def setup_overview():
    """Cria a tabela overview_items se ainda não existir."""
    conn = get_db()
    if not conn: return
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS overview_items (
                id           SERIAL PRIMARY KEY,
                area         VARCHAR(30)  NOT NULL,
                categoria    VARCHAR(30)  NOT NULL,
                descricao    TEXT         NOT NULL,
                prioridade   VARCHAR(20)  DEFAULT 'Media',
                status       VARCHAR(30)  DEFAULT 'aberto',
                responsavel  VARCHAR(100),
                data_prazo   DATE,
                usuario_id   INTEGER REFERENCES usuarios(id),
                criado_em    TIMESTAMP    DEFAULT NOW(),
                atualizado_em TIMESTAMP   DEFAULT NOW()
            )
        """)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_overview_area ON overview_items(area)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_overview_status ON overview_items(status)")
        conn.commit(); cur.close(); conn.close()
        print("✅ Tabela overview_items pronta")
    except Exception as e:
        print(f"Erro setup_overview: {e}")

setup_overview()

class OverviewItemModel(BaseModel):
    area:        str
    categoria:   str
    descricao:   str
    prioridade:  str = "Media"
    status:      str = "aberto"
    responsavel: Optional[str] = None
    data_prazo:  Optional[str] = None

@app.get("/api/overview")
def listar_overview(area: str = "", faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        if area and area in AREAS_VALIDAS:
            cur.execute("""
                SELECT id, area, categoria, descricao, prioridade, status,
                       responsavel, data_prazo::text, usuario_id, criado_em, atualizado_em
                FROM overview_items WHERE area=%s ORDER BY criado_em DESC
            """, (area,))
        else:
            cur.execute("""
                SELECT id, area, categoria, descricao, prioridade, status,
                       responsavel, data_prazo::text, usuario_id, criado_em, atualizado_em
                FROM overview_items ORDER BY area, criado_em DESC
            """)
        rows = cur.fetchall(); cur.close(); conn.close()
        return [
            {"id": r[0], "area": r[1], "categoria": r[2], "descricao": r[3],
             "prioridade": r[4], "status": r[5], "responsavel": r[6],
             "data_prazo": r[7], "usuario_id": r[8],
             "criado_em": str(r[9]), "atualizado_em": str(r[10])}
            for r in rows
        ]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/overview")
def criar_overview_item(item: OverviewItemModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    if item.area not in AREAS_VALIDAS:
        raise HTTPException(status_code=400, detail=f"Área inválida. Use: {AREAS_VALIDAS}")
    if item.categoria not in CATEGORIAS_VALIDAS:
        raise HTTPException(status_code=400, detail=f"Categoria inválida.")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO overview_items
                (area, categoria, descricao, prioridade, status, responsavel, data_prazo, usuario_id)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
        """, (item.area, item.categoria, item.descricao, item.prioridade, item.status,
              item.responsavel, item.data_prazo or None, sess["id"]))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/overview/{oid}")
def atualizar_overview_item(oid: int, item: OverviewItemModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    if item.area not in AREAS_VALIDAS:
        raise HTTPException(status_code=400, detail="Área inválida.")
    if item.categoria not in CATEGORIAS_VALIDAS:
        raise HTTPException(status_code=400, detail="Categoria inválida.")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        # Admin/gestor pode editar qualquer item; funcionário só o próprio
        if sess["perfil"] in ("admin", "gestor", "demo"):
            cur.execute("""
                UPDATE overview_items
                SET area=%s, categoria=%s, descricao=%s, prioridade=%s, status=%s,
                    responsavel=%s, data_prazo=%s, atualizado_em=NOW()
                WHERE id=%s
            """, (item.area, item.categoria, item.descricao, item.prioridade, item.status,
                  item.responsavel, item.data_prazo or None, oid))
        else:
            cur.execute("""
                UPDATE overview_items
                SET area=%s, categoria=%s, descricao=%s, prioridade=%s, status=%s,
                    responsavel=%s, data_prazo=%s, atualizado_em=NOW()
                WHERE id=%s AND usuario_id=%s
            """, (item.area, item.categoria, item.descricao, item.prioridade, item.status,
                  item.responsavel, item.data_prazo or None, oid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/overview/{oid}")
def deletar_overview_item(oid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        if sess["perfil"] in ("admin", "gestor", "demo"):
            cur.execute("DELETE FROM overview_items WHERE id=%s", (oid,))
        else:
            cur.execute("DELETE FROM overview_items WHERE id=%s AND usuario_id=%s", (oid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

# --- COMENTÁRIOS ---
class ComentarioModel(BaseModel):
    texto: str

@app.get("/api/tarefas/{tid}/comentarios")
def listar_comentarios(tid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("""SELECT c.id, c.texto, c.criado_em, u.nome
            FROM comentarios c JOIN usuarios u ON c.usuario_id = u.id
            WHERE c.tarefa_id = %s ORDER BY c.criado_em ASC""", (tid,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "texto": r[1], "criado_em": str(r[2]), "autor": r[3]} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/tarefas/{tid}/comentarios")
def criar_comentario(tid: int, c: ComentarioModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO comentarios (tarefa_id, usuario_id, texto) VALUES (%s,%s,%s) RETURNING id",
                    (tid, sess["id"], c.texto))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/comentarios/{cid}")
def deletar_comentario(cid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        if sess["perfil"] in ("admin", "gestor", "demo"):
            cur.execute("DELETE FROM comentarios WHERE id=%s", (cid,))
        else:
            cur.execute("DELETE FROM comentarios WHERE id=%s AND usuario_id=%s", (cid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

# --- RELATÓRIO POR CLIENTE ---
@app.get("/api/relatorio/{cliente}")
def get_relatorio(cliente: str, mes: str = "", faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        params_base = [cliente]
        filtro_mes = ""
        if mes:
            filtro_mes = "AND DATE_TRUNC('month', t.criado_em) = DATE_TRUNC('month', %s::date)"
            params_base.append(mes + "-01")

        cur.execute(
            "SELECT t.id, t.descricao, t.prioridade, t.status, t.segundos, t.criado_em, t.atualizado_em, u.nome "
            "FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id "
            "WHERE t.cliente = %s " + filtro_mes + " ORDER BY t.criado_em DESC",
            params_base)
        tarefas = cur.fetchall()

        cur.execute(
            "SELECT u.nome, COUNT(t.id), COALESCE(SUM(t.segundos),0) "
            "FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id "
            "WHERE t.cliente = %s " + filtro_mes + " GROUP BY u.nome ORDER BY SUM(t.segundos) DESC",
            params_base)
        por_func = cur.fetchall()

        cur.execute(
            "SELECT status, COUNT(*) FROM tarefas t WHERE t.cliente = %s " + filtro_mes + " GROUP BY status",
            params_base)
        status_counts = {r[0]: r[1] for r in cur.fetchall()}

        cur.execute(
            "SELECT COALESCE(SUM(segundos),0) FROM tarefas t WHERE t.cliente = %s " + filtro_mes,
            params_base)
        total_seg = cur.fetchone()[0]

        cur.close(); conn.close()
        status_map = {"concluido": "Concluído", "em_andamento": "Em Andamento", "aberto": "Aberto"}
        return {
            "cliente": cliente, "mes": mes,
            "resumo": {
                "total_tarefas": len(tarefas),
                "concluidas": status_counts.get("concluido", 0),
                "em_andamento": status_counts.get("em_andamento", 0),
                "abertas": status_counts.get("aberto", 0),
                "total_horas": round(total_seg / 3600, 1),
                "sla": round(status_counts.get("concluido", 0) / len(tarefas) * 100) if tarefas else 0
            },
            "por_funcionario": [{"nome": r[0], "tarefas": r[1], "horas": round(r[2]/3600, 1)} for r in por_func],
            "tarefas": [{"id": r[0], "descricao": r[1], "prioridade": r[2],
                "status": status_map.get(r[3], r[3]), "horas": round(r[4]/3600, 1),
                "minutos": (r[4] % 3600) // 60, "criado_em": str(r[5])[:10],
                "funcionario": r[7]} for r in tarefas]
        }
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/relatorio/{cliente}")
def relatorio_page(cliente: str): return FileResponse("static/relatorio.html")

@app.get("/apresentacao")
def apresentacao_page(): return FileResponse("static/apresentacao.html")

# --- NOTIFICAÇÕES ---
def criar_notificacao(conn, tipo: str, mensagem: str):
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO notificacoes (tipo, mensagem) VALUES (%s, %s)", (tipo, mensagem))
        # Mantém só as últimas 50
        cur.execute("DELETE FROM notificacoes WHERE id NOT IN (SELECT id FROM notificacoes ORDER BY criado_em DESC LIMIT 50)")
        cur.close()
    except:
        pass

@app.get("/api/notificacoes")
def get_notificacoes(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    if sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, tipo, mensagem, lida, criado_em FROM notificacoes ORDER BY criado_em DESC LIMIT 20")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "tipo": r[1], "mensagem": r[2], "lida": r[3], "criado_em": str(r[4])} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/notificacoes/marcar-lidas")
def marcar_lidas(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("UPDATE notificacoes SET lida = TRUE WHERE lida = FALSE")
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/notificacoes/nao-lidas")
def count_nao_lidas(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401)
    if sess["perfil"] not in ("admin", "gestor", "demo"): return {"count": 0}
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM notificacoes WHERE lida = FALSE")
        count = cur.fetchone()[0]
        cur.close(); conn.close()
        return {"count": count}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

# --- IA INSIGHTS ---
@app.post("/api/ia/insights")
def gerar_insights_ia(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"):
        raise HTTPException(status_code=403, detail="Acesso negado")

    groq_key = os.environ.get("GROQ_API_KEY", "")
    if not groq_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY não configurada")

    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")

    try:
        cur = conn.cursor()

        # Coleta dados para análise
        cur.execute("""
            SELECT u.nome,
                COUNT(t.id) as total,
                SUM(CASE WHEN t.status='aberto' THEN 1 ELSE 0 END) as abertos,
                SUM(CASE WHEN t.status='em_andamento' THEN 1 ELSE 0 END) as andamento,
                SUM(CASE WHEN t.status='concluido' THEN 1 ELSE 0 END) as concluidos,
                ROUND(AVG(t.segundos)/3600.0, 1) as media_horas
            FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id
            WHERE u.perfil = 'funcionario' AND u.ativo = TRUE
            GROUP BY u.nome ORDER BY total DESC
        """)
        por_func = cur.fetchall()

        cur.execute("""
            SELECT cliente, COUNT(*) as total,
                SUM(CASE WHEN status='aberto' THEN 1 ELSE 0 END) as abertos,
                SUM(CASE WHEN status='concluido' THEN 1 ELSE 0 END) as concluidos
            FROM tarefas GROUP BY cliente ORDER BY total DESC LIMIT 10
        """)
        por_cliente = cur.fetchall()

        cur.execute("""
            SELECT prioridade, COUNT(*) FROM tarefas WHERE status != 'concluido'
            GROUP BY prioridade
        """)
        por_prioridade = {r[0]: r[1] for r in cur.fetchall()}

        cur.execute("SELECT COUNT(*) FROM tarefas WHERE status='aberto' AND criado_em < NOW() - INTERVAL '7 days'")
        tickets_antigos = cur.fetchone()[0]

        cur.close(); conn.close()

        # Monta contexto para IA
        func_lines = "\n".join([
            f"- {r[0]}: {r[1]} tickets total, {r[2]} abertos, {r[3]} em andamento, {r[4]} concluídos, média {r[5]}h por ticket"
            for r in por_func
        ]) or "Nenhum dado"

        cliente_lines = "\n".join([
            f"- {r[0]}: {r[1]} tickets, {r[2]} abertos, {r[3]} concluídos"
            for r in por_cliente
        ]) or "Nenhum dado"

        criticos = por_prioridade.get('Critica', 0)
        altos = por_prioridade.get('Alta', 0)
        medios = por_prioridade.get('Media', 0)
        prompt = f"""Você é um assistente de operações da empresa Faiston. Analise APENAS os dados abaixo e gere exatamente 3 insights em português. Use somente os números fornecidos, não invente valores.

DADOS POR FUNCIONÁRIO (nome: total tickets, abertos, em andamento, concluídos):
{func_lines}

DADOS POR CLIENTE (cliente: total tickets, abertos, concluídos):
{cliente_lines}

RESUMO GERAL:
- Tickets abertos há mais de 7 dias (atrasados): {tickets_antigos}
- Tickets com prioridade CRÍTICA ainda abertos/em andamento: {criticos}
- Tickets com prioridade ALTA ainda abertos/em andamento: {altos}
- Tickets com prioridade MÉDIA ainda abertos/em andamento: {medios}

REGRAS:
- Use apenas os números acima, nunca some categorias diferentes
- Use os nomes reais dos funcionários
- Cada insight em uma linha, começando com emoji
- Máximo 130 caracteres por insight
- Foque nos problemas mais graves primeiro"""

        import json as _json, http.client, ssl
        body_json = _json.dumps({
            "model": "llama-3.1-8b-instant",
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": 400,
            "temperature": 0.7
        })
        ctx = ssl.create_default_context()
        conn = http.client.HTTPSConnection("api.groq.com", timeout=25, context=ctx)
        conn.request("POST", "/openai/v1/chat/completions", body=body_json, headers={
            "Authorization": f"Bearer {groq_key}",
            "Content-Type": "application/json",
            "User-Agent": "python-httpx/0.27",
            "Accept": "application/json",
        })
        resp = conn.getresponse()
        resp_body = resp.read().decode()
        conn.close()
        if resp.status != 200:
            print(f"[ia/insights] Groq {resp.status}: {resp_body}")
            raise HTTPException(status_code=500, detail=f"Groq {resp.status}: {resp_body}")
        result = _json.loads(resp_body)

        texto = result["choices"][0]["message"]["content"].strip()
        insights = [l.strip() for l in texto.split("\n") if l.strip()][:3]

        # Salva como notificações
        conn2 = get_db()
        if conn2:
            for insight in insights:
                criar_notificacao(conn2, "ia_insight", insight)
            conn2.commit()
            conn2.close()

        return {"insights": insights}

    except Exception as e:
        import traceback
        print(f"[ia/insights] ERRO: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

# --- HISTÓRICO COMPLETO ---
@app.get("/api/historico")
def get_historico(
    page: int = 1,
    por_pagina: int = 20,
    cliente: str = "",
    status: str = "",
    prioridade: str = "",
    funcionario: str = "",
    busca: str = "",
    data_inicio: str = "",
    data_fim: str = "",
    faiston_token: str = Cookie(None)
):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    if sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        conditions = ["1=1"]
        params = []
        if cliente:
            conditions.append("t.cliente = %s")
            params.append(cliente)
        if status:
            conditions.append("t.status = %s")
            params.append(status)
        if prioridade:
            conditions.append("t.prioridade = %s")
            params.append(prioridade)
        if funcionario:
            conditions.append("u.nome ILIKE %s")
            params.append(f"%{funcionario}%")
        if busca:
            conditions.append("t.descricao ILIKE %s")
            params.append(f"%{busca}%")
        if data_inicio:
            conditions.append("t.criado_em >= %s")
            params.append(data_inicio + " 00:00:00")
        if data_fim:
            conditions.append("t.criado_em <= %s")
            params.append(data_fim + " 23:59:59")

        where = "WHERE " + " AND ".join(conditions)
        offset = (page - 1) * por_pagina

        # Total de registros
        cur.execute(f"SELECT COUNT(*) FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id {where}", params)
        total = cur.fetchone()[0]

        # Tarefas paginadas
        cur.execute(f"""
            SELECT t.id, t.descricao, t.cliente, t.prioridade, t.status,
                   t.segundos, t.criado_em, t.atualizado_em, u.nome
            FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id
            {where} ORDER BY t.criado_em DESC
            LIMIT %s OFFSET %s
        """, params + [por_pagina, offset])
        rows = cur.fetchall()

        # Resumo do filtro atual
        cur.execute(f"""
            SELECT COALESCE(SUM(t.segundos),0),
                   SUM(CASE WHEN t.status='concluido' THEN 1 ELSE 0 END),
                   COUNT(*)
            FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id {where}
        """, params)
        resumo = cur.fetchone()

        cur.close(); conn.close()
        status_map = {"concluido": "Concluído", "em_andamento": "Em Andamento", "aberto": "Aberto"}
        return {
            "total": total,
            "pagina": page,
            "por_pagina": por_pagina,
            "total_paginas": max(1, -(-total // por_pagina)),
            "resumo": {
                "total_horas": round((resumo[0] or 0) / 3600, 1),
                "concluidas": resumo[1] or 0,
                "total": resumo[2] or 0
            },
            "tarefas": [{
                "id": r[0], "descricao": r[1], "cliente": r[2],
                "prioridade": r[3], "status": status_map.get(r[4], r[4]),
                "segundos": r[5], "criado_em": str(r[6])[:16],
                "atualizado_em": str(r[7])[:16], "funcionario": r[8]
            } for r in rows]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/historico")
def historico_page(): return FileResponse("static/historico.html")

# --- NOTAS PESSOAIS ---
class NotaModel(BaseModel):
    titulo: str
    texto: str

@app.get("/api/notas")
def listar_notas(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS notas (
                id SERIAL PRIMARY KEY,
                usuario_id INTEGER REFERENCES usuarios(id) ON DELETE CASCADE,
                titulo VARCHAR(200) NOT NULL,
                texto TEXT,
                criado_em TIMESTAMP DEFAULT NOW(),
                atualizado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit()
        cur.execute("""SELECT id, titulo, texto, criado_em, atualizado_em
            FROM notas WHERE usuario_id = %s ORDER BY atualizado_em DESC""",
            (sess["id"],))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "titulo": r[1], "texto": r[2],
                 "criado_em": str(r[3])[:16], "atualizado_em": str(r[4])[:16]} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/notas")
def criar_nota(n: NotaModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("INSERT INTO notas (usuario_id, titulo, texto) VALUES (%s,%s,%s) RETURNING id",
                    (sess["id"], n.titulo, n.texto))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/notas/{nid}")
def atualizar_nota(nid: int, n: NotaModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("UPDATE notas SET titulo=%s, texto=%s, atualizado_em=NOW() WHERE id=%s AND usuario_id=%s",
                    (n.titulo, n.texto, nid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/notas/{nid}")
def deletar_nota(nid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM notas WHERE id=%s AND usuario_id=%s", (nid, sess["id"]))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

# --- CLIENTES ---
@app.get("/api/clientes")
def listar_clientes(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401, detail="Não autenticado")
    conn = get_db()
    if not conn: raise HTTPException(status_code=500, detail="Banco offline")
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS clientes (
                id SERIAL PRIMARY KEY,
                nome VARCHAR(100) UNIQUE NOT NULL,
                contato VARCHAR(100),
                email VARCHAR(100),
                ativo BOOLEAN DEFAULT TRUE,
                criado_em TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS time VARCHAR(50) DEFAULT 'Projetos'")
        # Migrar clientes existentes das tarefas
        cur.execute("""
            INSERT INTO clientes (nome)
            SELECT DISTINCT cliente FROM tarefas
            WHERE cliente IS NOT NULL AND cliente != ''
            ON CONFLICT (nome) DO NOTHING
        """)
        conn.commit()
        if sess["perfil"] == "admin":
            cur.execute("SELECT id, nome, contato, email, ativo, criado_em, COALESCE(time,'Projetos') FROM clientes WHERE ativo=TRUE ORDER BY nome")
        else:
            cur.execute("SELECT id, nome, contato, email, ativo, criado_em, COALESCE(time,'Projetos') FROM clientes WHERE ativo=TRUE AND COALESCE(time,'Projetos')=%s ORDER BY nome", (sess.get("time","Projetos"),))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "nome": r[1], "contato": r[2], "email": r[3], "ativo": r[4], "criado_em": str(r[5])[:10], "time": r[6]} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

class ClienteModel(BaseModel):
    nome: str
    contato: str = ""
    email: str = ""
    ativo: bool = True
    time: str = "Projetos"

@app.post("/api/clientes")
def criar_cliente(c: ClienteModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        time_val = c.time if (sess["perfil"] == "admin" and c.time in TIMES_VALIDOS) else sess.get("time", "Projetos")
        cur.execute("INSERT INTO clientes (nome, contato, email, time) VALUES (%s,%s,%s,%s) RETURNING id",
                    (c.nome, c.contato, c.email, time_val))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except psycopg2.errors.UniqueViolation:
        raise HTTPException(status_code=400, detail="Cliente já existe")
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/clientes/{cid}")
def atualizar_cliente(cid: int, c: ClienteModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("UPDATE clientes SET nome=%s, contato=%s, email=%s, ativo=%s WHERE id=%s",
                    (c.nome, c.contato, c.email, c.ativo, cid))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/clientes/{cid}")
def deletar_cliente(cid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] != "admin": raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("UPDATE clientes SET ativo=FALSE WHERE id=%s", (cid,))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/ajuda")
def ajuda_page(): return FileResponse("static/ajuda.html")

@app.get("/clientes")
def clientes_page(): return FileResponse("static/clientes.html")

@app.get("/financeiro")
def financeiro_geral_page(): return FileResponse("static/financeiro-geral.html")

@app.get("/financeiro/{cid}")
def financeiro_page(cid: int): return FileResponse("static/financeiro.html")

@app.get("/api/financeiro/resumo")
def financeiro_resumo(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        _ensure_financeiro_tables(cur)
        conn.commit()
        cur.execute("""
            SELECT c.id, c.nome,
                   COALESCE(p_agg.num_projetos, 0) AS num_projetos,
                   COALESCE(p_agg.total_orcamento, 0) AS total_orcamento,
                   COALESCE(l_agg.total_gasto, 0) AS total_gasto
            FROM clientes c
            LEFT JOIN (
                SELECT cliente_id,
                       COUNT(*) AS num_projetos,
                       SUM(orcamento) AS total_orcamento
                FROM projetos
                WHERE ativo = TRUE
                GROUP BY cliente_id
            ) p_agg ON p_agg.cliente_id = c.id
            LEFT JOIN (
                SELECT p.cliente_id, SUM(l.valor) AS total_gasto
                FROM lancamentos l
                JOIN projetos p ON p.id = l.projeto_id AND p.ativo = TRUE
                GROUP BY p.cliente_id
            ) l_agg ON l_agg.cliente_id = c.id
            WHERE c.ativo = TRUE
            ORDER BY COALESCE(l_agg.total_gasto, 0) DESC, c.nome
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "nome": r[1], "num_projetos": r[2],
                 "orcamento": float(r[3]), "gasto": float(r[4])} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────
#  FINANCEIRO — Projetos e Lançamentos
# ─────────────────────────────────────────────

def _ensure_financeiro_tables(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS projetos (
            id SERIAL PRIMARY KEY,
            cliente_id INTEGER NOT NULL REFERENCES clientes(id),
            nome VARCHAR(100) NOT NULL,
            descricao TEXT DEFAULT '',
            orcamento NUMERIC(14,2) DEFAULT 0,
            ativo BOOLEAN DEFAULT TRUE,
            criado_em TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS lancamentos (
            id SERIAL PRIMARY KEY,
            projeto_id INTEGER NOT NULL REFERENCES projetos(id),
            descricao VARCHAR(200) NOT NULL,
            categoria VARCHAR(50) DEFAULT 'Outros',
            valor NUMERIC(14,2) NOT NULL,
            data_lancamento DATE DEFAULT CURRENT_DATE,
            criado_em TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("ALTER TABLE lancamentos ADD COLUMN IF NOT EXISTS localidade VARCHAR(150) DEFAULT ''")
    cur.execute("ALTER TABLE lancamentos ADD COLUMN IF NOT EXISTS tecnico VARCHAR(150) DEFAULT ''")
    cur.execute("ALTER TABLE projetos ADD COLUMN IF NOT EXISTS planilha_url TEXT DEFAULT ''")
    cur.execute("ALTER TABLE projetos ADD COLUMN IF NOT EXISTS planilha_mapeamento JSONB")
    cur.execute("ALTER TABLE projetos ADD COLUMN IF NOT EXISTS planilha_sync_em TIMESTAMP")
    cur.execute("ALTER TABLE projetos ADD COLUMN IF NOT EXISTS planilha_replace BOOLEAN DEFAULT FALSE")

class ProjetoModel(BaseModel):
    nome: str
    descricao: str = ""
    orcamento: float = 0.0

class LancamentoModel(BaseModel):
    descricao: str
    categoria: str = "Outros"
    valor: float
    data_lancamento: str = ""
    localidade: str = ""
    tecnico: str = ""

class LancamentoImportItem(BaseModel):
    descricao: str
    categoria: str = "Outros"
    valor: float
    data_lancamento: str = ""
    localidade: str = ""
    tecnico: str = ""

class ImportarProjetosPreviewBody(BaseModel):
    url: str

class ImportarProjetosBody(BaseModel):
    url: str
    col_cliente: str
    col_projeto: str
    col_orcamento: str = ""
    col_descricao: str = ""

class ImportarLancamentosBody(BaseModel):
    lancamentos: List[LancamentoImportItem]

class BaseImportacaoConfig(BaseModel):
    url: str
    col_cliente: str
    col_projeto: str
    col_orcamento: str = ""
    col_descricao: str = ""

@app.get("/api/config/base-importacao")
def get_base_importacao_config(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        import json as _json
        cur = conn.cursor()
        chave = f"base_importacao_{sess.get('time','Projetos')}"
        cur.execute("SELECT valor FROM configuracoes WHERE chave=%s", (chave,))
        row = cur.fetchone(); cur.close(); conn.close()
        if row:
            return _json.loads(row[0])
        return None
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/config/base-importacao")
def salvar_base_importacao_config(body: BaseImportacaoConfig, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        import json as _json
        cur = conn.cursor()
        chave = f"base_importacao_{sess.get('time','Projetos')}"
        valor = _json.dumps({"url": body.url, "col_cliente": body.col_cliente,
                             "col_projeto": body.col_projeto, "col_orcamento": body.col_orcamento,
                             "col_descricao": body.col_descricao})
        cur.execute("""INSERT INTO configuracoes (chave, valor, atualizado_em) VALUES (%s, %s, NOW())
                       ON CONFLICT (chave) DO UPDATE SET valor=EXCLUDED.valor, atualizado_em=NOW()""", (chave, valor))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/base-importacao/atualizar")
async def atualizar_base_importacao(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    global _OPENPYXL_OK, openpyxl
    if not _OPENPYXL_OK:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl as _ox; openpyxl = _ox; _OPENPYXL_OK = True
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        import json as _json
        cur = conn.cursor()
        time_sess = sess.get("time", "Projetos")
        chave = f"base_importacao_{time_sess}"
        cur.execute("SELECT valor FROM configuracoes WHERE chave=%s", (chave,))
        row = cur.fetchone()
        if not row: raise HTTPException(status_code=400, detail="Base não configurada para este time. Configure a URL primeiro.")
        cfg = _json.loads(row[0])
        _ensure_financeiro_tables(cur)
        content = _download_planilha(cfg["url"])
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers, rows = _parse_sheet(ws)
        headers_lower = [h.lower() for h in headers]
        def col_idx(name):
            if not name: return None
            try: return headers_lower.index(name.lower())
            except ValueError: return None
        idx_cliente = col_idx(cfg["col_cliente"])
        idx_projeto = col_idx(cfg["col_projeto"])
        idx_orc     = col_idx(cfg.get("col_orcamento", ""))
        idx_desc    = col_idx(cfg.get("col_descricao", ""))
        if idx_cliente is None or idx_projeto is None:
            raise HTTPException(status_code=400, detail=f"Colunas não encontradas na planilha. Colunas disponíveis: {', '.join(headers)}. Reconfigure a base.")
        # Filtra clientes apenas do time do usuário logado
        cur.execute("SELECT id, LOWER(nome) FROM clientes WHERE ativo=TRUE AND COALESCE(time,'Projetos')=%s", (time_sess,))
        clientes_db = {r[1].strip(): r[0] for r in cur.fetchall()}
        criados = ignorados = ja_existem = 0
        clientes_nao_encontrados = set()
        for row in rows:
            c_nome = row[idx_cliente].strip() if idx_cliente < len(row) else ""
            p_nome = row[idx_projeto].strip() if idx_projeto < len(row) else ""
            if not c_nome or not p_nome: ignorados += 1; continue
            cid = clientes_db.get(c_nome.lower())
            if not cid: clientes_nao_encontrados.add(c_nome); ignorados += 1; continue
            orc = 0.0
            if idx_orc is not None and idx_orc < len(row):
                try:
                    v = row[idx_orc].replace('R$','').replace('.','').replace(',','.').strip()
                    orc = float(v)
                except Exception: pass
            desc = row[idx_desc].strip() if idx_desc is not None and idx_desc < len(row) else ""
            cur.execute("SELECT id FROM projetos WHERE cliente_id=%s AND LOWER(nome)=%s AND ativo=TRUE", (cid, p_nome.lower()))
            if cur.fetchone(): ja_existem += 1; continue
            cur.execute("INSERT INTO projetos (cliente_id, nome, descricao, orcamento) VALUES (%s,%s,%s,%s)", (cid, p_nome, desc, orc))
            criados += 1
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "criados": criados, "ja_existem": ja_existem, "ignorados": ignorados,
                "clientes_nao_encontrados": sorted(clientes_nao_encontrados)}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/importar-projetos/preview")
async def importar_projetos_preview(body: ImportarProjetosPreviewBody, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    global _OPENPYXL_OK, openpyxl
    if not _OPENPYXL_OK:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl as _ox; openpyxl = _ox; _OPENPYXL_OK = True
    try:
        content = _download_planilha(body.url)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers, rows = _parse_sheet(ws)
        return {"headers": headers, "sample": rows[:5]}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/importar-projetos/executar")
async def importar_projetos_executar(body: ImportarProjetosBody, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    global _OPENPYXL_OK, openpyxl
    if not _OPENPYXL_OK:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl as _ox; openpyxl = _ox; _OPENPYXL_OK = True
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        content = _download_planilha(body.url)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers, rows = _parse_sheet(ws)

        headers_lower = [h.lower() for h in headers]
        def col_idx(name):
            if not name: return None
            try: return headers_lower.index(name.lower())
            except ValueError: return None

        idx_cliente  = col_idx(body.col_cliente)
        idx_projeto  = col_idx(body.col_projeto)
        idx_orc      = col_idx(body.col_orcamento) if body.col_orcamento else None
        idx_desc     = col_idx(body.col_descricao) if body.col_descricao else None

        if idx_cliente is None or idx_projeto is None:
            raise HTTPException(status_code=400, detail=f"Colunas não encontradas. Disponíveis: {', '.join(headers)}")

        cur = conn.cursor()
        _ensure_financeiro_tables(cur)
        time_sess = sess.get("time", "Projetos")
        cur.execute("SELECT id, LOWER(nome) FROM clientes WHERE ativo=TRUE AND COALESCE(time,'Projetos')=%s", (time_sess,))
        clientes_db = {r[1].strip(): r[0] for r in cur.fetchall()}

        criados = ignorados = ja_existem = 0
        clientes_nao_encontrados = set()
        for row in rows:
            c_nome = row[idx_cliente].strip() if idx_cliente < len(row) else ""
            p_nome = row[idx_projeto].strip() if idx_projeto < len(row) else ""
            if not c_nome or not p_nome:
                ignorados += 1; continue
            cid = clientes_db.get(c_nome.lower())
            if not cid:
                clientes_nao_encontrados.add(c_nome); ignorados += 1; continue
            orc = 0.0
            if idx_orc is not None and idx_orc < len(row):
                try:
                    v = row[idx_orc].replace('R$','').replace('.','').replace(',','.').strip()
                    orc = float(v)
                except Exception: pass
            desc = row[idx_desc].strip() if idx_desc is not None and idx_desc < len(row) else ""
            cur.execute("SELECT id FROM projetos WHERE cliente_id=%s AND LOWER(nome)=%s AND ativo=TRUE",
                        (cid, p_nome.lower()))
            if cur.fetchone():
                ja_existem += 1; continue
            cur.execute("INSERT INTO projetos (cliente_id, nome, descricao, orcamento) VALUES (%s,%s,%s,%s)",
                        (cid, p_nome, desc, orc))
            criados += 1

        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "criados": criados, "ja_existem": ja_existem, "ignorados": ignorados,
                "clientes_nao_encontrados": sorted(clientes_nao_encontrados)}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/todos-projetos")
def todos_projetos(faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        _ensure_financeiro_tables(cur); conn.commit()
        cur.execute("""
            SELECT p.id, p.nome, c.nome
            FROM projetos p
            JOIN clientes c ON c.id = p.cliente_id
            WHERE p.ativo = TRUE AND c.ativo = TRUE
            ORDER BY c.nome, p.nome
        """)
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"id": r[0], "nome": r[1], "cliente": r[2]} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/projetos-by-cliente")
def projetos_by_cliente_nome(nome: str = "", faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        _ensure_financeiro_tables(cur)
        conn.commit()
        if nome:
            cur.execute("SELECT id FROM clientes WHERE nome=%s AND ativo=TRUE", (nome,))
            row = cur.fetchone()
            if not row: return []
            cid = row[0]
            cur.execute("SELECT id, nome FROM projetos WHERE cliente_id=%s AND ativo=TRUE ORDER BY nome", (cid,))
        else:
            cur.execute("SELECT id, nome FROM projetos WHERE ativo=TRUE ORDER BY nome")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "nome": r[1]} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/clientes/{cid}/projetos")
def listar_projetos(cid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        _ensure_financeiro_tables(cur)
        conn.commit()
        cur.execute("""
            SELECT p.id, p.nome, p.descricao, p.orcamento, p.criado_em,
                   COALESCE(SUM(l.valor), 0) AS gasto,
                   COALESCE(p.planilha_url, '') AS planilha_url,
                   p.planilha_mapeamento,
                   p.planilha_sync_em,
                   COALESCE(p.planilha_replace, FALSE) AS planilha_replace
            FROM projetos p
            LEFT JOIN lancamentos l ON l.projeto_id = p.id
            WHERE p.cliente_id = %s AND p.ativo = TRUE
            GROUP BY p.id ORDER BY p.criado_em DESC
        """, (cid,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "nome": r[1], "descricao": r[2],
                 "orcamento": float(r[3]), "criado_em": str(r[4])[:10],
                 "gasto": float(r[5]),
                 "planilha_url": r[6] or "",
                 "planilha_mapeamento": r[7] or {},
                 "planilha_sync_em": str(r[8])[:16] if r[8] else None,
                 "planilha_replace": bool(r[9])} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/clientes/{cid}/analise-financeira")
def analise_financeira(cid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        _ensure_financeiro_tables(cur)
        conn.commit()

        # Gasto por localidade
        cur.execute("""
            SELECT COALESCE(NULLIF(TRIM(l.localidade),''), 'Não informado') AS loc,
                   SUM(l.valor) AS total
            FROM lancamentos l
            JOIN projetos p ON p.id = l.projeto_id
            WHERE p.cliente_id = %s AND p.ativo = TRUE
            GROUP BY loc ORDER BY total DESC LIMIT 10
        """, (cid,))
        loc_rows = cur.fetchall()
        total_loc = sum(r[1] for r in loc_rows) or 1
        por_localidade = [{"localidade": r[0], "gasto": float(r[1]),
                           "pct": round(float(r[1]) / total_loc * 100, 1)} for r in loc_rows]

        # Gasto por período (mês)
        cur.execute("""
            SELECT TO_CHAR(l.data_lancamento, 'YYYY-MM') AS periodo,
                   TO_CHAR(l.data_lancamento, 'Mon/YY') AS label,
                   SUM(l.valor) AS total
            FROM lancamentos l
            JOIN projetos p ON p.id = l.projeto_id
            WHERE p.cliente_id = %s AND p.ativo = TRUE
            GROUP BY periodo, label ORDER BY periodo
        """, (cid,))
        per_rows = cur.fetchall()
        max_per = max((float(r[2]) for r in per_rows), default=1)
        por_periodo = [{"periodo": r[0], "label": r[1], "gasto": float(r[2]),
                        "pct": round(float(r[2]) / max_per * 100, 1)} for r in per_rows]

        cur.close(); conn.close()
        return {"por_localidade": por_localidade, "por_periodo": por_periodo}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/clientes/{cid}/projetos")
def criar_projeto(cid: int, p: ProjetoModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        _ensure_financeiro_tables(cur)
        cur.execute("INSERT INTO projetos (cliente_id, nome, descricao, orcamento) VALUES (%s,%s,%s,%s) RETURNING id",
                    (cid, p.nome, p.descricao, p.orcamento))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/projetos/{pid}")
def atualizar_projeto(pid: int, p: ProjetoModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("UPDATE projetos SET nome=%s, descricao=%s, orcamento=%s WHERE id=%s",
                    (p.nome, p.descricao, p.orcamento, pid))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/projetos/{pid}")
def deletar_projeto(pid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("UPDATE projetos SET ativo=FALSE WHERE id=%s", (pid,))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/projetos/{pid}/lancamentos")
def listar_lancamentos(pid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess: raise HTTPException(status_code=401)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("""SELECT id, descricao, categoria, valor, data_lancamento, criado_em, localidade, tecnico
                       FROM lancamentos WHERE projeto_id=%s ORDER BY data_lancamento DESC, criado_em DESC""", (pid,))
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [{"id": r[0], "descricao": r[1], "categoria": r[2],
                 "valor": float(r[3]), "data": str(r[4]), "criado_em": str(r[5])[:10],
                 "localidade": r[6] or "", "tecnico": r[7] or ""} for r in rows]
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/projetos/{pid}/lancamentos")
def criar_lancamento(pid: int, l: LancamentoModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        data = l.data_lancamento or date.today().isoformat()
        cur.execute("INSERT INTO lancamentos (projeto_id, descricao, categoria, valor, data_lancamento, localidade, tecnico) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                    (pid, l.descricao, l.categoria, l.valor, data, l.localidade, l.tecnico))
        new_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "id": new_id}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/lancamentos/{lid}")
def deletar_lancamento(lid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM lancamentos WHERE id=%s", (lid,))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

def _detect_header_row(all_rows, max_search=10):
    """Return index of the row that best looks like a header (most non-empty text cells)."""
    best_idx, best_score = 0, 0
    for i, row in enumerate(all_rows[:max_search]):
        score = sum(1 for c in row if c is not None and isinstance(c, str) and c.strip())
        if score > best_score:
            best_score, best_idx = score, i
    return best_idx

def _parse_sheet(ws):
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []
    hi = _detect_header_row(all_rows)
    headers = [str(h).strip() if h is not None else "" for h in all_rows[hi]]
    rows = []
    for row in all_rows[hi+1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        rows.append([str(c).strip() if c is not None else "" for c in row])
    return headers, rows

@app.post("/api/projetos/{pid}/parse-planilha")
async def parse_planilha(pid: int, file: UploadFile = File(...),
                          sheet_name: str = "", faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    global _OPENPYXL_OK, openpyxl
    if not _OPENPYXL_OK:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl as _ox
        openpyxl = _ox
        _OPENPYXL_OK = True
    try:
        content = await file.read()
        filename = (file.filename or "").lower()

        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig", errors="replace")
            sample = text[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except Exception:
                dialect = csv.excel
            reader = csv.reader(io.StringIO(text), dialect)
            all_rows = list(reader)
            if not all_rows:
                return {"sheets": [], "selected_sheet": "", "headers": [], "rows": []}
            hi = _detect_header_row([[c for c in r] for r in all_rows])
            headers = [str(h).strip() for h in all_rows[hi]]
            rows = [[str(c).strip() for c in r]
                    for r in all_rows[hi+1:] if any(c.strip() for c in r)]
            return {"sheets": [], "selected_sheet": filename, "headers": headers, "rows": rows[:3000]}

        # Excel
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        # Discover sheets that have useful data (≥4 non-empty header cells)
        usable = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows_peek = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
            hi = _detect_header_row(rows_peek)
            hrow = rows_peek[hi] if rows_peek else []
            n_text = sum(1 for c in hrow if c and isinstance(c, str) and c.strip())
            if n_text >= 4:
                usable.append(sname)

        if not usable:
            raise ValueError("Nenhuma aba com dados tabulares encontrada")

        # Pick sheet
        chosen = sheet_name if sheet_name in usable else usable[0]
        # Prefer sheets with "VALOR FINAL"
        if not sheet_name:
            for s in usable:
                ws_tmp = wb[s]
                rows_tmp = list(ws_tmp.iter_rows(min_row=1, max_row=5, values_only=True))
                hi_tmp = _detect_header_row(rows_tmp)
                hrow_tmp = rows_tmp[hi_tmp] if rows_tmp else []
                if any('VALOR FINAL' in str(c) for c in hrow_tmp if c):
                    chosen = s
                    break

        ws = wb[chosen]
        headers, rows = _parse_sheet(ws)
        return {"sheets": usable, "selected_sheet": chosen, "headers": headers, "rows": rows[:3000]}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo: {str(e)}")

@app.post("/api/projetos/{pid}/importar-lancamentos")
def importar_lancamentos(pid: int, body: ImportarLancamentosBody, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        cur = conn.cursor()
        _ensure_financeiro_tables(cur)
        today = date.today().isoformat()
        count = 0
        for l in body.lancamentos:
            data = l.data_lancamento or today
            cur.execute(
                "INSERT INTO lancamentos (projeto_id, descricao, categoria, valor, data_lancamento, localidade, tecnico) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (pid, l.descricao[:200], l.categoria, l.valor, data, l.localidade[:150], l.tecnico[:150])
            )
            count += 1
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "importados": count}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────
#  PLANILHA ONLINE (OneDrive / SharePoint)
# ─────────────────────────────────────────────

class PlanilhaConfigModel(BaseModel):
    url: str
    mapeamento: dict
    replace_on_sync: bool = False

def _resolve_download_url(url: str) -> str:
    """Convert OneDrive/SharePoint share links to direct download URLs."""
    import urllib.parse
    u = url.strip()
    # SharePoint / OneDrive for Business
    if 'sharepoint.com' in u or 'onedrive.live.com' in u:
        sep = '&' if '?' in u else '?'
        return u + sep + 'download=1'
    # 1drv.ms short link — follow redirect then add download=1
    if '1drv.ms' in u or 'onedrive.com' in u:
        try:
            import requests as _req
            r = _req.head(u, timeout=15, allow_redirects=True)
            resolved = r.url
            sep = '&' if '?' in resolved else '?'
            return resolved + sep + 'download=1'
        except Exception:
            pass
    return u

def _download_planilha(url: str) -> bytes:
    import requests as _req
    dl_url = _resolve_download_url(url)
    r = _req.get(dl_url, timeout=60, allow_redirects=True,
                 headers={"User-Agent": "Mozilla/5.0"})
    if r.status_code != 200:
        raise ValueError(f"Erro ao baixar planilha (HTTP {r.status_code}). Verifique se o link permite acesso sem login.")
    if len(r.content) < 50:
        raise ValueError("Arquivo baixado está vazio. Verifique as permissões do link de compartilhamento.")
    return r.content

def _apply_mapeamento(headers, rows, mapeamento):
    """Convert raw sheet rows to LancamentoImportItem dicts using saved mapping."""
    from datetime import date as _date
    today = _date.today().isoformat()
    def col_idx(col_name):
        if not col_name:
            return None
        try:
            return headers.index(col_name)
        except ValueError:
            return None

    idx_desc  = col_idx(mapeamento.get("col_descricao"))
    idx_valor = col_idx(mapeamento.get("col_valor"))
    idx_data  = col_idx(mapeamento.get("col_data"))
    idx_cat   = col_idx(mapeamento.get("col_categoria"))
    idx_loc   = col_idx(mapeamento.get("col_localidade"))
    idx_tec   = col_idx(mapeamento.get("col_tecnico"))

    if idx_desc is None or idx_valor is None:
        raise ValueError("Mapeamento incompleto: coluna de descrição ou valor não encontrada nos cabeçalhos.")

    result = []
    for row in rows:
        def get(i): return row[i].strip() if i is not None and i < len(row) else ""
        raw_val = get(idx_valor).replace("R$","").replace(".","").replace(",",".").strip()
        try:
            val = float(raw_val)
        except Exception:
            continue
        if val == 0:
            continue
        result.append({
            "descricao":  get(idx_desc)[:200] or "—",
            "valor":       val,
            "data_lancamento": get(idx_data)[:10] if idx_data is not None else today,
            "categoria":  get(idx_cat)[:50] if idx_cat is not None else "Outros",
            "localidade": get(idx_loc)[:150] if idx_loc is not None else "",
            "tecnico":    get(idx_tec)[:150] if idx_tec is not None else "",
        })
    return result

@app.put("/api/projetos/{pid}/planilha-config")
def salvar_planilha_config(pid: int, body: PlanilhaConfigModel, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        import json
        cur = conn.cursor()
        _ensure_financeiro_tables(cur)
        cur.execute(
            "UPDATE projetos SET planilha_url=%s, planilha_mapeamento=%s, planilha_replace=%s WHERE id=%s",
            (body.url, json.dumps(body.mapeamento), body.replace_on_sync, pid)
        )
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/projetos/{pid}/testar-planilha")
def testar_planilha(pid: int, body: dict, faiston_token: str = Cookie(None)):
    """Download the file from stored/given URL and return its headers for mapping."""
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    global _OPENPYXL_OK, openpyxl
    if not _OPENPYXL_OK:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl as _ox; openpyxl = _ox; _OPENPYXL_OK = True
    url = body.get("url", "")
    if not url: raise HTTPException(status_code=400, detail="URL não informada")
    try:
        content = _download_planilha(url)
        import io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
        usable = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            rows_peek = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
            hi = _detect_header_row(rows_peek)
            hrow = rows_peek[hi] if rows_peek else []
            if sum(1 for c in hrow if c and isinstance(c, str) and c.strip()) >= 2:
                usable.append(sname)
        if not usable: raise ValueError("Nenhuma aba com dados encontrada")
        ws = wb[usable[0]]
        headers, rows = _parse_sheet(ws)
        return {"sheets": usable, "headers": headers, "sample": rows[:3]}
    except Exception as e: raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/projetos/{pid}/sincronizar")
def sincronizar_planilha(pid: int, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"): raise HTTPException(status_code=403)
    global _OPENPYXL_OK, openpyxl
    if not _OPENPYXL_OK:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl as _ox; openpyxl = _ox; _OPENPYXL_OK = True
    conn = get_db()
    if not conn: raise HTTPException(status_code=500)
    try:
        import io as _io
        cur = conn.cursor()
        _ensure_financeiro_tables(cur)
        cur.execute("SELECT planilha_url, planilha_mapeamento, planilha_replace FROM projetos WHERE id=%s AND ativo=TRUE", (pid,))
        row = cur.fetchone()
        if not row or not row[0]:
            raise HTTPException(status_code=400, detail="Planilha não configurada para este projeto")
        url, mapeamento, replace = row[0], row[1] or {}, bool(row[2])

        content = _download_planilha(url)
        wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
        sheet_name = mapeamento.get("sheet_name")
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        headers, rows = _parse_sheet(ws)
        lancamentos = _apply_mapeamento(headers, rows, mapeamento)

        if not lancamentos:
            raise HTTPException(status_code=400, detail="Nenhum lançamento encontrado com o mapeamento configurado")

        today = date.today().isoformat()
        # Sync sempre substitui — planilha é a fonte da verdade
        cur.execute("DELETE FROM lancamentos WHERE projeto_id=%s", (pid,))
        for l in lancamentos:
            cur.execute(
                "INSERT INTO lancamentos (projeto_id, descricao, categoria, valor, data_lancamento, localidade, tecnico) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (pid, l["descricao"], l["categoria"], l["valor"],
                 l["data_lancamento"] or today, l["localidade"], l["tecnico"])
            )
        cur.execute("UPDATE projetos SET planilha_sync_em=NOW() WHERE id=%s", (pid,))
        conn.commit(); cur.close(); conn.close()
        return {"sucesso": True, "importados": len(lancamentos), "replace": replace}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────
#  RELATÓRIO MENSAL
# ─────────────────────────────────────────────

def _coletar_dados_mes(ano: int, mes: int) -> dict:
    """Coleta KPIs, horas por funcionário e por cliente do mês especificado."""
    conn = get_db()
    if not conn:
        return {}
    try:
        cur = conn.cursor()
        inicio = date(ano, mes, 1).isoformat()
        fim    = date(ano, mes, monthrange(ano, mes)[1]).isoformat()

        # Total de tarefas e por status
        cur.execute("""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN status='concluido'    THEN 1 ELSE 0 END) AS concluidos,
                SUM(CASE WHEN status='em_andamento' THEN 1 ELSE 0 END) AS andamento,
                SUM(CASE WHEN status='aberto'       THEN 1 ELSE 0 END) AS abertos,
                COALESCE(SUM(segundos),0) AS total_seg
            FROM tarefas
            WHERE criado_em >= %s AND criado_em <= %s
        """, (inicio + " 00:00:00", fim + " 23:59:59"))
        row = cur.fetchone()
        total, concluidos, andamento, abertos, total_seg = row
        sla = round(concluidos / total * 100) if total else 0

        # Horas por funcionário (top 10)
        cur.execute("""
            SELECT u.nome, ROUND(SUM(t.segundos)/3600.0, 1) AS horas
            FROM tarefas t JOIN usuarios u ON t.usuario_id = u.id
            WHERE t.criado_em >= %s AND t.criado_em <= %s
            GROUP BY u.nome ORDER BY horas DESC LIMIT 10
        """, (inicio + " 00:00:00", fim + " 23:59:59"))
        por_func = [{"nome": r[0], "horas": float(r[1])} for r in cur.fetchall()]

        # Horas por cliente (top 8)
        cur.execute("""
            SELECT cliente, ROUND(SUM(segundos)/3600.0, 1) AS horas
            FROM tarefas
            WHERE criado_em >= %s AND criado_em <= %s AND cliente IS NOT NULL AND cliente != ''
            GROUP BY cliente ORDER BY horas DESC LIMIT 8
        """, (inicio + " 00:00:00", fim + " 23:59:59"))
        por_cliente = [{"nome": r[0], "horas": float(r[1])} for r in cur.fetchall()]

        cur.close(); conn.close()
        return {
            "ano": ano, "mes": mes,
            "total": total, "concluidos": concluidos,
            "andamento": andamento, "abertos": abertos,
            "total_horas": round(total_seg / 3600, 1),
            "sla": sla,
            "por_func": por_func,
            "por_cliente": por_cliente,
        }
    except Exception as e:
        print(f"Erro ao coletar dados relatório: {e}")
        return {}


MESES_PT = ["", "Janeiro","Fevereiro","Março","Abril","Maio","Junho",
            "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

def _gerar_html_relatorio(d: dict) -> str:
    if not d:
        return "<p>Sem dados disponíveis.</p>"

    mes_nome = MESES_PT[d["mes"]]
    max_h_func    = max((f["horas"] for f in d["por_func"]),    default=1)
    max_h_cliente = max((c["horas"] for c in d["por_cliente"]), default=1)

    def barra(pct, cor):
        return f'<div style="height:8px;border-radius:4px;background:#1E1E2E;margin-top:4px"><div style="height:8px;border-radius:4px;background:{cor};width:{pct}%"></div></div>'

    linhas_func = "".join(f"""
        <tr>
          <td style="padding:10px 0;border-bottom:1px solid #1E1E2E;color:#C8CBE0;font-size:14px">{f['nome'].split()[0]}</td>
          <td style="padding:10px 0;border-bottom:1px solid #1E1E2E;width:55%">
            {barra(round(f['horas']/max_h_func*100), 'linear-gradient(90deg,#5B2EE0,#06D7E6)')}
          </td>
          <td style="padding:10px 0;border-bottom:1px solid #1E1E2E;text-align:right;color:#06D7E6;font-weight:700;font-size:14px">{f['horas']}h</td>
        </tr>""" for f in d["por_func"])

    linhas_cliente = "".join(f"""
        <tr>
          <td style="padding:10px 0;border-bottom:1px solid #1E1E2E;color:#C8CBE0;font-size:14px">{c['nome']}</td>
          <td style="padding:10px 0;border-bottom:1px solid #1E1E2E;width:50%">
            {barra(round(c['horas']/max_h_cliente*100), 'linear-gradient(90deg,#B826C9,#EC4899)')}
          </td>
          <td style="padding:10px 0;border-bottom:1px solid #1E1E2E;text-align:right;color:#EC4899;font-weight:700;font-size:14px">{c['horas']}h</td>
        </tr>""" for c in d["por_cliente"])

    sla_cor = "#06D7E6" if d["sla"] >= 92 else "#EC4899"

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Relatório Mensal · Faiston OPS · {mes_nome} {d['ano']}</title></head>
<body style="margin:0;padding:0;background:#07070F;font-family:'Helvetica Neue',Arial,sans-serif;color:#E2E8F0">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#07070F;padding:32px 0">
<tr><td align="center">
<table width="620" cellpadding="0" cellspacing="0" style="background:#0D0D1A;border:1px solid #1E1E2E;border-radius:16px;overflow:hidden">

  <!-- HEADER -->
  <tr><td style="background:linear-gradient(135deg,#5B2EE0,#B826C9,#06D7E6);padding:2px 0"></td></tr>
  <tr><td style="padding:32px 40px 24px;border-bottom:1px solid #1E1E2E">
    <table width="100%"><tr>
      <td>
        <div style="font-size:11px;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:#5B2EE0;margin-bottom:6px">Relatório Operacional</div>
        <div style="font-size:28px;font-weight:800;color:#FFFFFF;letter-spacing:-0.5px">FAISTON <span style="background:linear-gradient(90deg,#5B2EE0,#06D7E6);-webkit-background-clip:text;-webkit-text-fill-color:transparent">OPS</span></div>
        <div style="font-size:14px;color:#5E647A;margin-top:4px">{mes_nome} {d['ano']} · Gerado automaticamente</div>
      </td>
      <td align="right" style="vertical-align:top">
        <div style="display:inline-block;background:rgba(6,215,230,0.08);border:1px solid rgba(6,215,230,0.25);border-radius:8px;padding:8px 16px;font-size:12px;font-weight:700;color:#06D7E6">SLA {d['sla']}% <span style="color:{'#06D7E6' if d['sla']>=92 else '#EC4899'}">{'✓ Meta' if d['sla']>=92 else '✗ Abaixo'}</span></div>
      </td>
    </tr></table>
  </td></tr>

  <!-- KPIs -->
  <tr><td style="padding:28px 40px;border-bottom:1px solid #1E1E2E">
    <table width="100%" cellspacing="0" cellpadding="0"><tr>
      <td align="center" style="padding:16px;background:#12121F;border-radius:12px;border:1px solid #1E1E2E">
        <div style="font-size:32px;font-weight:900;color:#06D7E6">{d['total_horas']}h</div>
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#5E647A;margin-top:4px">Horas Apontadas</div>
      </td>
      <td width="12"></td>
      <td align="center" style="padding:16px;background:#12121F;border-radius:12px;border:1px solid #1E1E2E">
        <div style="font-size:32px;font-weight:900;color:#5B2EE0">{d['total']}</div>
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#5E647A;margin-top:4px">Total de Tickets</div>
      </td>
      <td width="12"></td>
      <td align="center" style="padding:16px;background:#12121F;border-radius:12px;border:1px solid #1E1E2E">
        <div style="font-size:32px;font-weight:900;color:#B826C9">{d['concluidos']}</div>
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#5E647A;margin-top:4px">Concluídos</div>
      </td>
      <td width="12"></td>
      <td align="center" style="padding:16px;background:#12121F;border-radius:12px;border:1px solid #1E1E2E">
        <div style="font-size:32px;font-weight:900;color:{sla_cor}">{d['sla']}%</div>
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#5E647A;margin-top:4px">SLA</div>
      </td>
    </tr></table>
  </td></tr>

  <!-- EQUIPE -->
  <tr><td style="padding:28px 40px;border-bottom:1px solid #1E1E2E">
    <div style="font-size:10px;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:#5E647A;margin-bottom:16px">Equipe · Horas por Funcionário</div>
    <table width="100%" cellspacing="0" cellpadding="0">{linhas_func}</table>
  </td></tr>

  <!-- CLIENTES -->
  <tr><td style="padding:28px 40px;border-bottom:1px solid #1E1E2E">
    <div style="font-size:10px;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:#5E647A;margin-bottom:16px">Clientes · Distribuição de Horas</div>
    <table width="100%" cellspacing="0" cellpadding="0">{linhas_cliente}</table>
  </td></tr>

  <!-- FOOTER -->
  <tr><td style="padding:24px 40px;text-align:center">
    <div style="font-size:11px;color:#3A3A55">Relatório gerado automaticamente pelo <strong style="color:#5B2EE0">Faiston OPS</strong> · rafael.libel@faiston.com</div>
  </td></tr>

</table>
</td></tr></table>
</body></html>"""


def _enviar_email(html: str, mes_nome: str, ano: int):
    email_to = os.environ.get("EMAIL_RELATORIO", "rafael.libel@faiston.com")
    subject  = f"Relatório Faiston OPS · {mes_nome} {ano}"

    # ── Brevo (API HTTP — funciona no Railway) ───────────────────────
    import urllib.request, json as _json
    brevo_key = os.environ.get("BREVO_API_KEY", "")
    email_user = os.environ.get("EMAIL_USER", "")
    if brevo_key and email_user:
        payload = _json.dumps({
            "sender": {"name": "Faiston OPS", "email": email_user},
            "to": [{"email": email_to}],
            "subject": subject,
            "htmlContent": html,
        }).encode()
        req = urllib.request.Request(
            "https://api.brevo.com/v3/smtp/email",
            data=payload,
            headers={"api-key": brevo_key, "Content-Type": "application/json"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = _json.loads(resp.read())
        print(f"Brevo OK — id {result.get('messageId')} — {mes_nome}/{ano} → {email_to}")
        return

    # ── Fallback SMTP (Gmail) ─────────────────────────────────────────
    smtp_host = os.environ.get("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASS", "")
    if not smtp_user or not smtp_pass:
        raise ValueError("Configure RESEND_API_KEY (recomendado) ou SMTP_USER + SMTP_PASS")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = f"Faiston OPS <{smtp_user}>"
    msg["To"]      = email_to
    msg.attach(MIMEText(html, "html", "utf-8"))

    ctx = ssl.create_default_context()
    try:
        with smtplib.SMTP_SSL(smtp_host, 465, context=ctx, timeout=15) as s:
            s.login(smtp_user, smtp_pass)
            s.sendmail(smtp_user, email_to, msg.as_string())
    except Exception as e1:
        print(f"SMTP_SSL 465 falhou: {e1} — tentando STARTTLS 587")
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as s:
            s.ehlo(); s.starttls(context=ctx); s.ehlo()
            s.login(smtp_user, smtp_pass)
            s.sendmail(smtp_user, email_to, msg.as_string())
    print(f"SMTP OK — Relatório {mes_nome}/{ano} enviado para {email_to}")


def _job_relatorio_mensal():
    hoje = date.today()
    # Mês anterior
    primeiro_do_mes = hoje.replace(day=1)
    mes_ant = primeiro_do_mes - timedelta(days=1)
    d = _coletar_dados_mes(mes_ant.year, mes_ant.month)
    if not d:
        print("Relatório mensal: sem dados para enviar")
        return
    html = _gerar_html_relatorio(d)
    try:
        _enviar_email(html, MESES_PT[mes_ant.month], mes_ant.year)
    except Exception as e:
        print(f"Erro ao enviar relatório: {e}")


# Inicia agendador
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    _scheduler = BackgroundScheduler(timezone="America/Sao_Paulo")
    _scheduler.add_job(_job_relatorio_mensal, "cron", day=1, hour=8, minute=0)
    _scheduler.start()
    print("APScheduler iniciado — relatório agendado para dia 1 de cada mês às 08h")
except ImportError:
    print("APScheduler não instalado — relatórios automáticos desativados. Instale com: pip install apscheduler")


# ─── Endpoints de relatório ───

@app.get("/api/relatorio-mensal/preview", response_class=HTMLResponse)
def preview_relatorio(mes: int = 0, ano: int = 0, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"):
        raise HTTPException(status_code=403)
    hoje = date.today()
    if not mes:
        # Tenta mês anterior; se vazio, usa mês atual
        ref = (hoje.replace(day=1) - timedelta(days=1))
        d = _coletar_dados_mes(ref.year, ref.month)
        if not d or d.get("total", 0) == 0:
            mes, ano = hoje.month, hoje.year
        else:
            mes, ano = ref.month, ref.year
    if not ano:
        ano = hoje.year
    d = _coletar_dados_mes(ano, mes)
    return HTMLResponse(_gerar_html_relatorio(d))


@app.post("/api/relatorio-mensal/enviar")
def enviar_relatorio_manual(mes: int = 0, ano: int = 0, faiston_token: str = Cookie(None)):
    sess = get_session(faiston_token)
    if not sess or sess["perfil"] not in ("admin", "gestor", "demo"):
        raise HTTPException(status_code=403)
    hoje = date.today()
    if not mes:
        # Tenta mês anterior; se vazio, usa mês atual
        ref = (hoje.replace(day=1) - timedelta(days=1))
        d_test = _coletar_dados_mes(ref.year, ref.month)
        if not d_test or d_test.get("total", 0) == 0:
            mes, ano = hoje.month, hoje.year
        else:
            mes, ano = ref.month, ref.year
    if not ano:
        ano = hoje.year
    d = _coletar_dados_mes(ano, mes)
    if not d:
        raise HTTPException(status_code=404, detail="Sem dados para o período")
    html = _gerar_html_relatorio(d)
    try:
        _enviar_email(html, MESES_PT[mes], ano)
        return {"sucesso": True, "mensagem": f"Relatório {MESES_PT[mes]}/{ano} enviado"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))